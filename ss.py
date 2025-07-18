from flask import Flask, render_template, request, redirect, url_for, send_from_directory, session, flash
import pandas as pd
import os 
import re 
import csv
import sqlite3
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'your_secret_key'

DATABASE = 'data/feedback.db'
USERS_FILE = 'data/users.csv'
EXCEL_FILE = 'data/staff_feedback.xlsx' 

headers = [
    'ID', 'S/N', 'Activity', 'Department', 'Division', 'Start Date', 'Date of Last Update', 'Name', 'Work Done', 
    'Status', 'Recommendation', 'Approval from ECOP (if any)'
]

def get_current_week_sheet_name():
    """Returns sheet name in format 'YYYY-WW' based on current date"""
    today = datetime.now()
    year, week_num, _ = today.isocalendar()
    return f"{year}-W{week_num:02d}"

def init_db():
    if not os.path.exists('data'):
        os.makedirs('data')
        
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS feedback (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            department TEXT,
            division TEXT,
            activity TEXT,
            work_done TEXT,
            start_date TEXT,
            status TEXT,
            recommendation TEXT,
            ecop_approval TEXT,
            week TEXT,
            last_update TEXT
        )
    ''')

    conn.commit()
    conn.close()

def init_excel():
    if not os.path.exists('data'):
        os.makedirs('data')
    
    current_sheet_name = get_current_week_sheet_name()
    
    if not os.path.isfile(EXCEL_FILE): 
        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
            pd.DataFrame().to_excel(writer, sheet_name=current_sheet_name, index=False)
        
        format_sheet(current_sheet_name)
    else:
        wb = load_workbook(EXCEL_FILE)
        if current_sheet_name not in wb.sheetnames:
            wb.create_sheet(current_sheet_name)
            wb.save(EXCEL_FILE)
            format_sheet(current_sheet_name)
        wb.close()

def format_sheet(sheet_name):
    """Format a sheet with the required structure"""
    wb = load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]
    
    ws.column_dimensions['A'].hidden = True
    
    for col in ['C', 'D', 'E', 'H', 'I', 'K', 'L']:
        ws.column_dimensions[col].width = 48.00

    for col in ['F', 'G', 'J']:
        ws.column_dimensions[col].width = 18.00
    
    header_font = Font(bold=True, color='FFFFFF', size=14)
    header_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = header_font
        cell.fill = header_fill

    wb.save(EXCEL_FILE)

def get_next_available_row(sheet_name):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]
    row = 2
    while ws.cell(row=row, column=1).value is not None:
        row += 1
    wb.close()
    return row

def save_to_excel(entry):
    current_sheet_name = get_current_week_sheet_name()
    wb = load_workbook(EXCEL_FILE)
    ws = wb[current_sheet_name]

    row = 2
    while ws.cell(row=row, column=1).value is not None:
        row += 1

    sn = row - 1 

    def insert_line_breaks(text, max_length=60):
        if not text or not isinstance(text, str):
            return text
        words = text.split(' ')
        lines = []
        current_line = ""
        for word in words:
            if len(current_line) + len(word) + 1 <= max_length:
                current_line += (" " + word) if current_line else word
            else:
                if current_line:
                    lines.append(current_line)
                current_line = word
        if current_line:
            lines.append(current_line)
        return '\n'.join(lines)

    activity = insert_line_breaks(entry['Activity'])
    work_done = insert_line_breaks(entry['Work Done'])
    recommendation = insert_line_breaks(entry['Recommendation'])
    approval = insert_line_breaks(entry['Approval from ECOP (if any)'])

    ws.cell(row=row, column=1, value=entry['ID'])
    ws.cell(row=row, column=2, value=sn)
    ws.cell(row=row, column=3, value=activity)
    ws.cell(row=row, column=4, value=entry['Department'])
    ws.cell(row=row, column=5, value=entry['Division'])
    ws.cell(row=row, column=6, value=entry['Start Date'])
    ws.cell(row=row, column=7, value=None)
    ws.cell(row=row, column=8, value=entry['Name'])
    ws.cell(row=row, column=9, value=work_done)
    ws.cell(row=row, column=10, value=entry['Status'])
    ws.cell(row=row, column=11, value=recommendation)
    ws.cell(row=row, column=12, value=approval)

    for col in [3, 9, 11, 12]:
        ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)

    wb.save(EXCEL_FILE)
    wb.close()


def read_all_entries():
    wb = load_workbook(EXCEL_FILE)
    entries = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        row = 2
        while ws.cell(row=row, column=1).value is not None:
            entry = {
                'ID': ws.cell(row=row, column=1).value,
                'Activity': ws.cell(row=row, column=3).value,
                'Department': ws.cell(row=row, column=4).value,
                'Division': ws.cell(row=row, column=5).value,
                'Start Date': ws.cell(row=row, column=6).value,
                'Last Update': ws.cell(row=row, column=7).value,
                'Name': ws.cell(row=row, column=8).value,
                'Work Done': ws.cell(row=row, column=9).value,
                'Status': ws.cell(row=row, column=10).value,
                'Recommendation': ws.cell(row=row, column=11).value,
                'Approval from ECOP (if any)': ws.cell(row=row, column=12).value,
                'Week': sheet_name
            }
            entries.append(entry)
            row += 1

    wb.close()
    return entries


def update_entry(entry_id, updated_data):
    """Find and update an entry by ID across all sheets"""
    wb = load_workbook(EXCEL_FILE)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        row = 2

        while ws.cell(row=row, column=1).value is not None:
            if ws.cell(row=row, column=1).value == entry_id:
                ws.cell(row=row, column=3, value=updated_data['Activity'])
                ws.cell(row=row, column=5, value=updated_data['Start Date'])
                ws.cell(row=row, column=6, value=updated_data.get('last_update'))
                ws.cell(row=row, column=8, value=updated_data['Work Done'])
                ws.cell(row=row, column=9, value=updated_data['Status'])
                ws.cell(row=row, column=10, value=updated_data['Recommendation'])
                ws.cell(row=row, column=11, value=updated_data['Approval from ECOP (if any)'])

                for col in [3, 9, 11, 12]:
                    ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)

                wb.save(EXCEL_FILE)
                wb.close()
                return True
            row += 1

    wb.close()
    return False


@app.route('/form', methods=['GET', 'POST'])
def form_page():
    if not session.get('user_logged_in'):
        return redirect(url_for('user_login'))
    
    init_excel()
    if request.method == 'POST':
        names = request.form['name']
        division = request.form['division']
        department=request.form['department']
        comments = request.form['comment']

        work_done_list = request.form.getlist('work_done[]')
        date_list = request.form.getlist('date[]')
        status_list = request.form.getlist('status[]')
        Activity_list = request.form.getlist('Activity[]')
        recommendation_list = request.form.getlist('recommendation[]')

        all_entries = read_all_entries()
        existing_max_id = max([entry['ID'] for entry in all_entries]) if all_entries else 0

        conn = sqlite3.connect(DATABASE)
        c = conn.cursor()

        current_week = get_current_week_sheet_name()
        for i in range(len(work_done_list)):
            entry = {
                'ID': existing_max_id + i + 1,
                'Name': names,
                'Department': department,
                'Division': division,
                'Work Done': work_done_list[i],
                'Start Date': date_list[i],
                'Status': status_list[i],
                'Activity': Activity_list[i],
                'Recommendation': recommendation_list[i],
                'Approval from ECOP (if any)': comments
            }

            c.execute('''INSERT INTO feedback 
                (name, department, division, activity, work_done, start_date, status, recommendation, ecop_approval, week, last_update) 
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', 
                (entry['Name'], entry['Department'], entry['Division'], entry['Activity'], entry['Work Done'], entry['Start Date'],
                entry['Status'], entry['Recommendation'], entry['Approval from ECOP (if any)'], current_week, None))

            entry['ID'] = c.lastrowid
            save_to_excel(entry)

        conn.commit()
        conn.close()

        return redirect(url_for('submissions'))

    return render_template('form_multi.html')

@app.route('/submissions', methods=['GET'])
def submissions():
    q = request.args.get('q', '').lower()
    is_admin = session.get('is_admin', False)
    all_entries = read_all_entries()

    df = pd.DataFrame(all_entries)

    if is_admin:
        filtered = df  
    elif q:
        mask = (
            df['Name'].str.lower().str.contains(q, na=False) |
            df['Division'].str.lower().str.contains(q, na=False) |
            df['Activity'].str.lower().str.contains(q, na=False)
        )
        filtered = df[mask]
    else:
        filtered = df.iloc[0:0]

    data = filtered.to_dict(orient='records')
    return render_template('submissions.html', data=data, is_admin=is_admin)

@app.route('/download')
def download():
    if not session.get('is_admin'):
        return "Unauthorized", 403

    if not os.path.exists(EXCEL_FILE):
        return "Report file not found. Please submit some data first.", 404

    directory = os.path.abspath(os.path.dirname(EXCEL_FILE))
    filename = os.path.basename(EXCEL_FILE)
    current_sheet = get_current_week_sheet_name()

    return send_from_directory(
        directory=directory,
        path=filename,
        as_attachment=True,
        download_name=f"DRMD_Weekly_Report_{current_sheet}.xlsx"
    )

@app.route('/edit/<int:entry_id>', methods=['GET', 'POST'])
def edit(entry_id):
    init_excel()
    entries = read_all_entries()
    entry = next((e for e in entries if e['ID'] == entry_id), None)

    if not entry:
        return "Entry not found.", 404

    if request.method == 'POST':
        updated_data = {
            'Work Done': request.form['work_done'],
            'Start Date': request.form['date'],
            'Status': request.form['status'],
            'Activity': request.form['Activity'],
            'Recommendation': request.form['recommendation'],
            'Approval from ECOP (if any)': request.form['comment'],
            'last_update': request.form.get('last_update')
        }
        update_entry(entry_id, updated_data)
        return redirect(url_for('submissions'))

    return render_template('edit.html', entry=entry)

def register_user(username, password, email):
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, newline='') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row['username'] == username:
                    return 'username_taken'
                if row['email'] == email:
                    return 'email_taken'

    with open(USERS_FILE, 'a', newline='') as f:
        writer = csv.writer(f)
        if os.path.getsize(USERS_FILE) == 0:
            writer.writerow(['username', 'password', 'email']) 
        writer.writerow([username, password, email])
    return 'success'


def validate_user(username, password):
    if not os.path.exists(USERS_FILE):
        return False
    with open(USERS_FILE, newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row['username'] == username and row['password'] == password:
                return True
    return False

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        if username == 'admin' and password == 'adminpass123':
            session['is_admin'] = True
            session.pop('division', None)
            return redirect(url_for('submissions'))
        else:
            flash('Invalid admin username or password', 'danger')
    return render_template('login.html')

@app.route('/', methods=['GET', 'POST'])
def user_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        if validate_user(username, password):
            session['user_logged_in'] = True
            session['username'] = username
            return redirect(url_for('form_page'))
        else:
            flash('Invalid username or password', 'error')

    return render_template('user_login.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        confirm_password = request.form['confirm_password']
        email = request.form['email']

        if password != confirm_password:
            flash('Passwords do not match.', 'error')
            return render_template('signup.html')

        if len(password) <= 4 or not re.search(r'\d', password):
            flash('Password must be more than 4 characters and contain at least one number.', 'error')
            return render_template('signup.html')

        result = register_user(username, password, email)
        if result == 'success':
            flash('Account created successfully. Please log in.', 'success')
            return redirect(url_for('user_login'))
        elif result == 'username_taken':
            flash('Username already taken. Please choose another.', 'error')
        elif result == 'email_taken':
            flash('Email already used. Please use a different email.', 'error')

    return render_template('signup.html')


@app.route('/logout')
def logout():
    session.pop('is_admin', None)
    return redirect(url_for('submissions'))


if __name__ == '__main__':
    init_db()
    init_excel()
    app.run(debug=True)
